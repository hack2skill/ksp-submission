import pandas as pd
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import math


def process_workbook(file):
    wb = xl.open(file)
    string = ""
    sheet1 = wb['Sheet1']
    for i in range(2, sheet1.max_row + 1):
        cell = sheet1.cell(i, 2)  # <Cell 'Sheet1'.B3144> returns the object which represent the cell in the Excel sheet
        string = string + f" {cell.value}"  # cell value returns the value corresponding to that cell

    list1 = string.split(" ")
    list1.remove(list1[0])
    l1 = []
    for i in range(0, len(list1)):
        number = list1[i]
        l1.append(number[0])

    dict1 = {'noofrep': l1}
    dataset = pd.DataFrame(dict1)
    result = dataset.pivot_table(columns=['noofrep'], aggfunc='size')
    probability(result, sheet1)
    f_name = input("enter file name to save as:")
    wb.save(f"C:\\Users\\Dhiraj\\Videos\\BL_USpopulation\\{f_name}.xlsx")
    print(f"we can know conclude your result by watching file at C:/Users/Dhiraj/Videos/BL_USpopulation/{f_name}.xlsx :)")


def probability(dframe, sheet1):
    sheet1.cell(1, 3).value = 'counts'
    sheet1.cell(1, 4).value = 'observed'
    sheet1.cell(1, 5).value = 'Benfords\'s'
    l1 = dframe.values  # returns the list of values of specified column
    #print(l1)
    for i in range(2, 11):
        sheet1.cell(i, 3).value = i - 1
        sheet1.cell(i, 4).value = (l1[i - 2] / len(df)) * 100
        sheet1.cell(i, 5).value = math.log((1 + 1 / (i - 1)), 10) * 100

    chart1 = BarChart()
    values1 = Reference(sheet1, min_col=4, min_row=2, max_row=9, max_col=4)
    sheet1.cell(1, 9).value = "GRAPH"
    chart1.add_data(values1)
    values2 = Reference(sheet1, min_col=5, min_row=2, max_row=9, max_col=5)
    chart1.add_data(values2)
    sheet1.add_chart(chart1, 'h2')


path = input("enter the absolute file path :(use '/' instead of '\' between folder names :")
# path -> C:/Users/Dhiraj/Videos/BL_USpopulation/USpopulation.xlsx -> use in this pattern .
df = pd.read_excel(path)
process_workbook(path)
