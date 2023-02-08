
import pandas as pd
import re
import openpyxl


def upi():
    df = pd.read_excel("transactions_2023_02_04_21_23_25.xlsx")
    data_list = df[df.columns[2]].tolist()
    date = df[df.columns[1]].tolist()
    reference = df[df.columns[3]].tolist()
    debit = df[df.columns[4]].tolist()
    credit = df[df.columns[5]].tolist()
    balance = df[df.columns[6]].tolist()

    transactionType = []
    TransactionAnalyser = []

    transaction_categories = {
        "UPI": {"count": 0, "details_map": {}},
        "ATM": {"count": 0, "details_map": {}},
        "debit card": {"count": 0, "details_map": {}},
        "NEFT": {"count": 0, "details_map": {}},
        "IMPS": {"count": 0, "details_map": {}},
    }

    for i in range(0, len(data_list)):
        data = data_list[i]

        # UPI
        if "UPI" in data:
            first_index = data.index("/")
            second_index = data.index("/", first_index + 1)

            if "/" in data[second_index + 1:]:
                start_index = data.index("/", second_index + 1) + 1
            else:
                start_index = second_index + 1

            end_index = data.rindex("/")
            result = data[start_index:end_index]
            transaction_categories["UPI"]["count"] += 1
            if result in transaction_categories["UPI"]["details_map"]:
                transaction_categories["UPI"]["details_map"][result] += 1
            else:
                transaction_categories["UPI"]["details_map"][result] = 1
            TransactionAnalyser.append(result)

        # ATM
        elif "ATM" in data:
            result = data.split("-")[1].strip()
            transaction_categories["ATM"]["count"] += 1
            if result in transaction_categories["ATM"]["details_map"]:
                transaction_categories["ATM"]["details_map"][result] += 1
            else:
                transaction_categories["ATM"]["details_map"][result] = 1
            TransactionAnalyser.append(result)

        # Debit Card
        elif "debit card" in data:
            result = data.split("-")[1].strip()
            transaction_categories["debit card"]["count"] += 1
            if result in transaction_categories["debit card"]["details_map"]:
                transaction_categories["debit card"]["details_map"][result] += 1
            else:
                transaction_categories["debit card"]["details_map"][result] = 1
            TransactionAnalyser.append(result)

            # NEFT
        elif "NEFT" in data:
            result = data.split("-")[1].strip()
            transaction_categories["NEFT"]["count"] += 1
            if result in transaction_categories["NEFT"]["details_map"]:
                transaction_categories["NEFT"]["details_map"][result] += 1
            else:
                transaction_categories["NEFT"]["details_map"][result] = 1
            TransactionAnalyser.append(result)

        # IMPS
        elif "IMPS" in data:
            result = data.split("-")[1].strip()
            transaction_categories["IMPS"]["count"] += 1
            if result in transaction_categories["IMPS"]["details_map"]:
                transaction_categories["IMPS"]["details_map"][result] += 1
            else:
                transaction_categories["IMPS"]["details_map"][result] = 1
            TransactionAnalyser.append(result)

    headers = ["Date", "TransactionType", "Details",
               "Reference", "Debit", "Credit", "Balance"]
    data = [date, transactionType, TransactionAnalyser,
            reference, debit, credit, balance]

    # Create a new workbook object
    workbook = openpyxl.Workbook()

    # Create a new sheet in the workbook
    sheet = workbook.active
    sheet.title = "My Sheet"

    # Fill headers
    for i, header in enumerate(headers):
        sheet.cell(row=1, column=i+1, value=header)

    # Fill data
    for k, row_data in enumerate(data):
        for j, cell_data in enumerate(row_data):
            sheet.cell(row=j+2, column=k+1, value=cell_data)

    # Save the workbook to a file
    workbook.save("new_file3.xlsx")
